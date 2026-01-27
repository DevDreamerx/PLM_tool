# -*- coding: utf-8 -*-
import re
from datetime import datetime, date
from difflib import SequenceMatcher

from openpyxl import load_workbook


def _normalize(text):
    if text is None:
        return ""
    value = str(text).strip().lower()
    value = value.replace("\n", " ").replace("\r", " ")
    value = re.sub(r"[\s_\-\/\\:：()（）\[\]【】]+", "", value)
    return value


def _similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()


def _clean_value(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if text in {"——", "--", "-", "—"}:
        return ""
    return text


class ExcelImporter:
    """Excel 导入工具（自动表头识别）"""

    def __init__(self):
        self._multi_fields = {"change_description", "change_order"}
        self._label_fields = {"change_description", "change_order"}
        self._synonyms = {
            "product_code": [
                "产品代号", "代号", "产品代码", "物料编码", "产品编号", "产品号",
                "产品型号"
            ],
            "product_name": [
                "产品名称", "名称", "品名", "物料名称"
            ],
            "batch_number": [
                "批次编号", "批次号", "批次", "批号"
            ],
            "model": [
                "型号", "产品型号", "机型", "型号类别",
                "所属机型"
            ],
            "drawing_number": [
                "图号", "图纸号", "图样号", "图纸编号"
            ],
            "drawing_version": [
                "图纸版本", "图纸版次", "图号版本", "图纸版本号"
            ],
            "software_version": [
                "软件版本", "软件版本号", "软件号", "软件构建", "软件build", "sw版本", "软件版本构建"
            ],
            "firmware_version": [
                "固件版本", "固件版本号", "固件号", "固件构建", "固件build", "fw版本", "固件版本构建"
            ],
            "hardware_config": [
                "硬件配置", "硬件说明", "硬件参数", "硬件配置说明"
            ],
            "req_baseline": [
                "需求基线", "功能基线", "需求版本", "需求基线版本"
            ],
            "icd_version": [
                "接口基线", "接口版本", "接口文档版本", "icd", "icd版本"
            ],
            "bom_version": [
                "bom版本", "物料清单版本", "bom版次", "物料表版本"
            ],
            "pcb_version": [
                "pcb版本", "板卡版本", "pcb版次", "板号"
            ],
            "hw_serial": [
                "硬件序列号", "序列号", "硬件sn", "sn", "s/n"
            ],
            "production_batch": [
                "生产批次", "生产批号", "制造批次"
            ],
            "test_status": [
                "测试状态", "试验状态", "验证状态", "测试进度"
            ],
            "qual_status": [
                "合格状态", "放行状态", "鉴定状态", "合格评审"
            ],
            "change_order": [
                "更改单号", "变更单号", "变更编号", "更改单号", "ecn", "eco",
                "协调单号", "更改建议单号", "更改单号/技术通知单号/工艺更改单号"
            ],
            "change_description": [
                "更改内容", "变更内容", "变更说明", "更改说明", "变更摘要",
                "更改理由", "更改原因", "更改类别", "处理意见", "更改建议单涉及图样/文件", "涉及更改图样",
                "更改人", "需落实产品编号", "已落实情况", "未落实产品编号", "工艺更改落实情况", "备注",
                "所属阶段"
            ],
            "effective_date": [
                "生效日期", "生效时间", "变更日期", "日期"
            ],
        }
        self._synonyms_norm = {
            key: [_normalize(item) for item in items]
            for key, items in self._synonyms.items()
        }

    def guess_header_row(self, ws, max_scan=8):
        best_row = 1
        best_score = -1
        for row_idx in range(1, min(ws.max_row, max_scan) + 1):
            values = [_normalize(cell.value) for cell in ws[row_idx]]
            score = 0
            for val in values:
                if self.match_field(val):
                    score += 1
            if score > best_score:
                best_score = score
                best_row = row_idx
        return best_row

    def match_field(self, header_value):
        if not header_value:
            return None
        # 特殊规则：生产批次优先
        if "生产" in header_value and "批次" in header_value:
            return "production_batch"
        if "落实" in header_value and "产品编号" in header_value:
            return "change_description"
        for field, synonyms in self._synonyms_norm.items():
            for syn in synonyms:
                if not syn:
                    continue
                if header_value == syn or syn in header_value or header_value in syn:
                    return field
        # 模糊匹配
        best_field = None
        best_score = 0.0
        for field, synonyms in self._synonyms_norm.items():
            for syn in synonyms:
                score = _similarity(header_value, syn)
                if score > best_score:
                    best_score = score
                    best_field = field
        if best_score >= 0.72:
            return best_field
        return None

    def build_mapping(self, ws, header_row):
        mapping = {}
        for idx, cell in enumerate(ws[header_row], start=1):
            header_raw = "" if cell.value is None else str(cell.value).strip()
            header = _normalize(header_raw)
            if not header:
                continue
            field = self.match_field(header)
            if not field:
                continue
            if field in self._multi_fields:
                mapping.setdefault(field, []).append((header_raw, idx))
            elif field not in mapping:
                mapping[field] = idx
        return mapping

    def parse(self, file_path, sheet_name=None):
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        header_row = self.guess_header_row(ws)
        mapping = self.build_mapping(ws, header_row)

        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            record = {}
            for field, col_info in mapping.items():
                if isinstance(col_info, list):
                    combined = []
                    for header_raw, col_idx in col_info:
                        value = row[col_idx - 1] if col_idx - 1 < len(row) else None
                        value = _clean_value(value)
                        if value:
                            if field in self._label_fields and header_raw:
                                combined.append(f"{header_raw}:{value}")
                            else:
                                combined.append(value)
                    record[field] = "; ".join(combined)
                else:
                    value = row[col_info - 1] if col_info - 1 < len(row) else None
                    record[field] = _clean_value(value)
            rows.append(record)
        return {
            "header_row": header_row,
            "mapping": mapping,
            "rows": rows,
        }
