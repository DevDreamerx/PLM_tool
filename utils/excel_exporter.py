# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

class ExcelExporter:
    """Excel导出工具类"""
    
    @staticmethod
    def export_products(data_list, output_dir="."):
        """
        导出产品数据到Excel
        
        Args:
            data_list: 产品数据列表，每项包含product和tech_status信息
            output_dir: 输出目录
            
        Returns:
            str: 生成的文件路径
        """
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "技术状态数据"
        
        # 定义表头
        headers = [
            "ID", "产品代号", "产品名称", "批次编号", "所属型号",
            "图号", "图纸版本", "软件版本", "固件版本", "硬件配置",
            "更改单号", "更改内容", "生效日期", "状态", "创建时间"
        ]
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row_num, data in enumerate(data_list, 2):
            # 产品基本信息
            ws.cell(row=row_num, column=1, value=data.get('id', ''))
            ws.cell(row=row_num, column=2, value=data.get('product_code', ''))
            ws.cell(row=row_num, column=3, value=data.get('product_name', ''))
            ws.cell(row=row_num, column=4, value=data.get('batch_number', ''))
            ws.cell(row=row_num, column=5, value=data.get('model', ''))
            
            # 技术状态信息
            ws.cell(row=row_num, column=6, value=data.get('drawing_number', ''))
            ws.cell(row=row_num, column=7, value=data.get('drawing_version', ''))
            ws.cell(row=row_num, column=8, value=data.get('software_version', ''))
            ws.cell(row=row_num, column=9, value=data.get('firmware_version', ''))
            ws.cell(row=row_num, column=10, value=data.get('hardware_config', ''))
            ws.cell(row=row_num, column=11, value=data.get('change_order', ''))
            ws.cell(row=row_num, column=12, value=data.get('change_description', ''))
            ws.cell(row=row_num, column=13, value=data.get('effective_date', ''))
            
            # 状态和时间
            status_text = "草稿" if data.get('status') == 'draft' else "正式" if data.get('status') == 'active' else "已删除"
            ws.cell(row=row_num, column=14, value=status_text)
            ws.cell(row=row_num, column=15, value=data.get('created_at', ''))
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"技术状态数据_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # 保存文件
        wb.save(filepath)
        return filepath
