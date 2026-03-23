# -*- coding: utf-8 -*-
import os
import platform

import matplotlib
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QFileDialog,
    QFrame,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QVBoxLayout,
    QWidget,
)

from db.database import DatabaseManager
from ui.theme import THEME, scale_px


def setup_matplotlib_fonts():
    if platform.system() == "Windows":
        fonts = ["Microsoft YaHei", "SimHei", "Arial Unicode MS"]
    elif platform.system() == "Darwin":
        fonts = ["Arial Unicode MS", "PingFang HK", "Heiti TC", "sans-serif"]
    else:
        fonts = ["DejaVu Sans", "Bitstream Vera Sans", "sans-serif"]

    matplotlib.rcParams["font.sans-serif"] = fonts
    matplotlib.rcParams["axes.unicode_minus"] = False


setup_matplotlib_fonts()
matplotlib.use("Qt5Agg")


class ReportWidget(QWidget):
    """数据报表界面"""

    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self._stat_labels = []
        self.init_ui()

    def init_ui(self):
        self._apply_report_styles()

        root_layout = QVBoxLayout()
        root_layout.setContentsMargins(0, 0, 0, 0)

        content = QWidget()
        content.setObjectName("ReportContainer")
        content.setMaximumWidth(1260)
        content.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(18)

        title = QLabel("技术状态报表")
        title.setObjectName("PageTitle")
        main_layout.addWidget(title)

        intro = QLabel("从总体规模、型号分布和最近变更活动三个维度观察当前技术状态。")
        intro.setObjectName("ReportIntro")
        intro.setWordWrap(True)
        main_layout.addWidget(intro)

        hero = QFrame()
        hero.setObjectName("HeroPanel")
        hero_layout = QHBoxLayout()
        hero_layout.setContentsMargins(24, 22, 24, 22)
        hero_layout.setSpacing(18)

        hero_text_wrap = QVBoxLayout()
        hero_text_wrap.setSpacing(8)
        hero_title = QLabel("技术状态总览")
        hero_title.setObjectName("HeroTitle")
        hero_desc = QLabel("在一个页面里同时看清总量、正式状态占比、草稿积压和最新变更动态。")
        hero_desc.setObjectName("HeroDesc")
        hero_desc.setWordWrap(True)
        hero_text_wrap.addWidget(hero_title)
        hero_text_wrap.addWidget(hero_desc)
        hero_text_wrap.addStretch()

        hero_action_wrap = QHBoxLayout()
        hero_action_wrap.setSpacing(12)
        self.btn_refresh = QPushButton("刷新报表")
        self.btn_refresh.setObjectName("GhostButton")
        self.btn_refresh.setMinimumSize(112, 42)
        self.btn_refresh.clicked.connect(self.refresh_data)
        self.btn_export = QPushButton("导出全部数据")
        self.btn_export.setObjectName("PrimaryButton")
        self.btn_export.setMinimumSize(132, 42)
        self.btn_export.clicked.connect(self.export_all_data)
        hero_action_wrap.addWidget(self.btn_refresh)
        hero_action_wrap.addWidget(self.btn_export)
        hero_action_wrap.addStretch()

        hero_text_wrap.addLayout(hero_action_wrap)
        hero_layout.addLayout(hero_text_wrap, 3)

        accent_panel = QFrame()
        accent_panel.setObjectName("HeroAccent")
        accent_layout = QVBoxLayout()
        accent_layout.setContentsMargins(20, 18, 20, 18)
        accent_layout.setSpacing(6)
        accent_tag = QLabel("Report")
        accent_tag.setObjectName("AccentTag")
        accent_value = QLabel("变更管理")
        accent_value.setObjectName("AccentValue")
        accent_hint = QLabel("围绕型号分布与最新活动组织")
        accent_hint.setObjectName("AccentHint")
        accent_layout.addWidget(accent_tag)
        accent_layout.addWidget(accent_value)
        accent_layout.addWidget(accent_hint)
        accent_layout.addStretch()
        accent_panel.setLayout(accent_layout)
        hero_layout.addWidget(accent_panel, 1)

        hero.setLayout(hero_layout)
        main_layout.addWidget(hero)

        stats_layout = QHBoxLayout()
        stats_layout.setSpacing(14)
        self.total_card = self.create_stat_card("总记录数", "0", "全部产品", "#315f8d")
        self.active_card = self.create_stat_card("正式记录", "0", "已发布状态", "#2f7d62")
        self.draft_card = self.create_stat_card("草稿数", "0", "待完善数据", "#b7791f")
        stats_layout.addWidget(self.total_card)
        stats_layout.addWidget(self.active_card)
        stats_layout.addWidget(self.draft_card)
        main_layout.addLayout(stats_layout)

        insight_layout = QHBoxLayout()
        insight_layout.setSpacing(18)

        chart_group = QGroupBox("型号分布")
        chart_group.setObjectName("ReportCard")
        chart_layout = QVBoxLayout()
        chart_layout.setContentsMargins(18, 18, 18, 18)
        chart_layout.setSpacing(12)

        chart_hint = QLabel("对比不同型号当前存量，便于快速识别集中分布。")
        chart_hint.setObjectName("ReportHint")
        self.figure = Figure(figsize=(8, 4))
        self.figure.set_facecolor("#f8fbfd")
        self.canvas = FigureCanvas(self.figure)
        chart_layout.addWidget(chart_hint)
        chart_layout.addWidget(self.canvas)
        chart_group.setLayout(chart_layout)
        insight_layout.addWidget(chart_group, 3)

        changes_group = QGroupBox("最近变更活动")
        changes_group.setObjectName("ReportCard")
        changes_layout = QVBoxLayout()
        changes_layout.setContentsMargins(18, 18, 18, 18)
        changes_layout.setSpacing(12)

        changes_hint = QLabel("默认展示最近 8 条变更记录。")
        changes_hint.setObjectName("ReportHint")
        self.changes_list = QListWidget()
        self.changes_list.setObjectName("ChangesList")
        self.changes_list.setSpacing(8)
        self.changes_list.setMinimumWidth(340)
        changes_layout.addWidget(changes_hint)
        changes_layout.addWidget(self.changes_list)
        changes_group.setLayout(changes_layout)
        insight_layout.addWidget(changes_group, 2)

        main_layout.addLayout(insight_layout)

        content.setLayout(main_layout)

        centered = QWidget()
        centered_layout = QHBoxLayout()
        centered_layout.setContentsMargins(18, 18, 18, 24)
        centered_layout.addStretch()
        centered_layout.addWidget(content)
        centered_layout.addStretch()
        centered.setLayout(centered_layout)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QScrollArea.NoFrame)
        scroll_area.setWidget(centered)
        root_layout.addWidget(scroll_area)
        self.setLayout(root_layout)

        self.refresh_data()

    def _apply_report_styles(self):
        self.setStyleSheet(
            f"""
            QWidget#ReportContainer {{
                background: transparent;
            }}
            QLabel#ReportIntro {{
                color: {THEME['text_muted']};
                font-size: 14px;
                padding: 0 4px 2px 4px;
            }}
            QFrame#HeroPanel {{
                background: qlineargradient(
                    x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f7fafc,
                    stop:0.55 #edf4fa,
                    stop:1 #e7f0f8
                );
                border: 1px solid #d8e2ec;
                border-radius: 22px;
            }}
            QLabel#HeroTitle {{
                color: {THEME['text']};
                font-size: 22px;
                font-weight: 700;
            }}
            QLabel#HeroDesc {{
                color: {THEME['text_muted']};
                font-size: 13px;
                line-height: 1.5;
            }}
            QFrame#HeroAccent {{
                background: #315f8d;
                border-radius: 18px;
            }}
            QLabel#AccentTag {{
                color: rgba(255, 255, 255, 0.72);
                font-size: 11px;
                font-weight: 700;
                letter-spacing: 1px;
            }}
            QLabel#AccentValue {{
                color: #ffffff;
                font-size: 24px;
                font-weight: 700;
            }}
            QLabel#AccentHint {{
                color: rgba(255, 255, 255, 0.8);
                font-size: 12px;
            }}
            QGroupBox#StatCard, QGroupBox#ReportCard {{
                border: 1px solid #dbe4ee;
                border-radius: 18px;
                margin-top: 10px;
                background: #ffffff;
                padding: 6px 0 0 0;
            }}
            QGroupBox#ReportCard::title, QGroupBox#StatCard::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 14px;
                padding: 0 8px;
                color: {THEME['text_muted']};
                font-size: 13px;
                font-weight: 700;
                background: {THEME['bg_app']};
                border-radius: 8px;
            }}
            QLabel#ReportHint {{
                color: {THEME['text_muted']};
                font-size: 12px;
            }}
            QPushButton#GhostButton {{
                min-height: 42px;
                color: {THEME['text']};
                border: 1px solid #c7d5e6;
                border-radius: 12px;
                background: #ffffff;
                font-weight: 600;
            }}
            QPushButton#GhostButton:hover {{
                background: #f4f8fc;
                border-color: #9bb4d1;
            }}
            QPushButton#PrimaryButton {{
                min-height: 42px;
                color: #ffffff;
                background: #315f8d;
                border: 1px solid #315f8d;
                border-radius: 12px;
                font-weight: 700;
            }}
            QPushButton#PrimaryButton:hover {{
                background: #294f76;
                border-color: #294f76;
            }}
            QListWidget#ChangesList {{
                background: transparent;
                border: none;
                padding: 0;
            }}
            QListWidget#ChangesList::item {{
                border: 1px solid #e4eaf1;
                background: #f9fbfd;
                border-radius: 14px;
                padding: 12px;
                margin: 0;
            }}
            QListWidget#ChangesList::item:selected {{
                background: #eef4f8;
                color: {THEME['text']};
            }}
            """
        )

    def create_stat_card(self, title, value, subtitle, accent_color):
        card = QGroupBox(title)
        card.setObjectName("StatCard")
        layout = QVBoxLayout()
        layout.setContentsMargins(18, 20, 18, 18)
        layout.setSpacing(8)

        value_label = QLabel(value)
        value_label.setObjectName("value")
        value_label.setStyleSheet(
            f"font-size: {scale_px(30)}px; font-weight: 700; color: {accent_color};"
        )

        subtitle_label = QLabel(subtitle)
        subtitle_label.setStyleSheet(
            f"font-size: {scale_px(12)}px; color: {THEME['text_muted']};"
        )

        layout.addWidget(value_label)
        layout.addWidget(subtitle_label)
        layout.addStretch()
        card.setLayout(layout)
        self._stat_labels.append((value_label, subtitle_label, accent_color))
        return card

    def apply_font_scale(self, scale):
        for value_label, subtitle_label, accent_color in getattr(self, "_stat_labels", []):
            value_label.setStyleSheet(
                f"font-size: {scale_px(30, scale)}px; font-weight: 700; color: {accent_color};"
            )
            subtitle_label.setStyleSheet(
                f"font-size: {scale_px(12, scale)}px; color: {THEME['text_muted']};"
            )

    def refresh_data(self):
        stats = self.db.get_statistics()
        self.total_card.findChild(QLabel, "value").setText(str(stats["total_count"]))
        self.active_card.findChild(QLabel, "value").setText(str(stats["active_count"]))
        self.draft_card.findChild(QLabel, "value").setText(str(stats["draft_count"]))

        self.update_chart()
        self.update_recent_changes()

    def update_chart(self):
        distribution = self.db.get_model_distribution()

        self.figure.clear()
        ax = self.figure.add_subplot(111)
        ax.set_facecolor("#f8fbfd")

        if not distribution:
            ax.text(
                0.5,
                0.5,
                "暂无型号分布数据",
                ha="center",
                va="center",
                fontsize=13,
                color=THEME["text_muted"],
                transform=ax.transAxes,
            )
            ax.axis("off")
            self.figure.tight_layout()
            self.canvas.draw()
            return

        models = [item[0] or "未分类" for item in distribution]
        counts = [item[1] for item in distribution]
        bars = ax.barh(models, counts, color=["#5f85ad", "#7b9ec1", "#98b6d3", "#b4ccdf", "#d0e1ec"][:len(models)])

        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_visible(False)
        ax.spines["bottom"].set_color("#d5dfeb")
        ax.tick_params(axis="x", colors=THEME["text_muted"], labelsize=10)
        ax.tick_params(axis="y", colors=THEME["text"], labelsize=11, length=0)
        ax.grid(axis="x", linestyle="--", linewidth=0.8, color="#e3eaf1")
        ax.set_axisbelow(True)
        ax.set_xlabel("数量", fontsize=11, color=THEME["text_muted"])
        ax.set_title("产品型号分布", fontsize=14, fontweight="bold", color=THEME["text"], loc="left", pad=10)
        ax.invert_yaxis()

        max_count = max(counts) if counts else 0
        ax.set_xlim(0, max_count * 1.18 if max_count else 1)

        for bar, count in zip(bars, counts):
            ax.text(
                bar.get_width() + max_count * 0.03,
                bar.get_y() + bar.get_height() / 2,
                str(count),
                va="center",
                ha="left",
                fontsize=10,
                color=THEME["text"],
            )

        self.figure.tight_layout(pad=1.4)
        self.canvas.draw()

    def update_recent_changes(self):
        self.changes_list.clear()
        changes = self.db.get_recent_changes(limit=8)

        if not changes:
            empty_item = QListWidgetItem("暂无变更活动")
            self.changes_list.addItem(empty_item)
            return

        for item in changes:
            created_at = item.get("created_at", "")
            time_text = created_at[5:16] if len(created_at) >= 16 else created_at
            content = item.get("change_content", "") or "未填写变更内容"
            if len(content) > 42:
                content = content[:42] + "..."
            line = (
                f"{time_text}\n"
                f"{item.get('product_code', '-')}"
                f"  {item.get('product_name', '')}\n"
                f"{content}\n"
                f"操作人：{item.get('operator', '系统')}"
            )
            self.changes_list.addItem(QListWidgetItem(line))

    def export_all_data(self):
        try:
            data = self.db.get_products_with_tech_status()
            if not data:
                QMessageBox.warning(self, "提示", "没有可导出的数据")
                return

            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存Excel文件", "", "Excel Files (*.xlsx)"
            )
            if not file_path:
                return
            if not file_path.endswith(".xlsx"):
                file_path += ".xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "技术状态数据"

            headers = [
                "ID", "产品代号", "产品名称", "批次编号", "所属型号",
                "图号", "图纸版本", "软件版本", "固件版本", "硬件配置",
                "需求基线", "接口基线", "BOM版本", "PCB版本", "硬件序列号",
                "生产批次", "测试状态", "合格状态",
                "更改单号", "更改内容", "生效日期", "状态", "创建时间",
            ]

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="315F8D", end_color="315F8D", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row_num, item in enumerate(data, 2):
                ws.cell(row=row_num, column=1, value=item.get("id", ""))
                ws.cell(row=row_num, column=2, value=item.get("product_code", ""))
                ws.cell(row=row_num, column=3, value=item.get("product_name", ""))
                ws.cell(row=row_num, column=4, value=item.get("batch_number", ""))
                ws.cell(row=row_num, column=5, value=item.get("model", ""))
                ws.cell(row=row_num, column=6, value=item.get("drawing_number", ""))
                ws.cell(row=row_num, column=7, value=item.get("drawing_version", ""))
                ws.cell(row=row_num, column=8, value=item.get("software_version", ""))
                ws.cell(row=row_num, column=9, value=item.get("firmware_version", ""))
                ws.cell(row=row_num, column=10, value=item.get("hardware_config", ""))
                ws.cell(row=row_num, column=11, value=item.get("req_baseline", ""))
                ws.cell(row=row_num, column=12, value=item.get("icd_version", ""))
                ws.cell(row=row_num, column=13, value=item.get("bom_version", ""))
                ws.cell(row=row_num, column=14, value=item.get("pcb_version", ""))
                ws.cell(row=row_num, column=15, value=item.get("hw_serial", ""))
                ws.cell(row=row_num, column=16, value=item.get("production_batch", ""))
                ws.cell(row=row_num, column=17, value=item.get("test_status", ""))
                ws.cell(row=row_num, column=18, value=item.get("qual_status", ""))
                ws.cell(row=row_num, column=19, value=item.get("change_order", ""))
                ws.cell(row=row_num, column=20, value=item.get("change_description", ""))
                ws.cell(row=row_num, column=21, value=item.get("effective_date", ""))
                status_text = "草稿" if item.get("status") == "draft" else "正式"
                ws.cell(row=row_num, column=22, value=status_text)
                ws.cell(row=row_num, column=23, value=item.get("created_at", ""))

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            wb.save(file_path)
            QMessageBox.information(self, "成功", f"数据已导出到:\n{file_path}")
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", f"导出过程中发生错误:\n{exc}")
